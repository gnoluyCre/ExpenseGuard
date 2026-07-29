"""Deterministic XLSX encoding for immutable F4 report snapshots."""

from __future__ import annotations

import json
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any, Final, cast
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.core.errors import ExpenseGuardError
from app.core.policies.canonical import canonical_sha256
from app.core.reports.models import ReportSnapshot

XLSX_TEMPLATE_VERSION: Final = "report-xlsx-v1"
MAX_CELL_CHARACTERS: Final = 32_767
SHEET_NAMES: Final = ("摘要", "关注项", "原始行证据", "解析错误", "制度快照")
type CellValue = str | bool | int | float | date | datetime | None

SUMMARY_HEADERS: Final = ("metric", "value")
SUMMARY_METRICS: Final = (
    "file_version_id",
    "filename",
    "revision_no",
    "source_content_sha256",
    "validation_run_id",
    "report_run_id",
    "report_fingerprint",
    "ruleset_fingerprint",
    "mapping_version_id",
    "policy_fingerprint",
    "binding_fingerprint",
    "report_template_version",
    "export_template_version",
    "generated_at",
    "stored_row_count",
    "validated_row_count",
    "flagged_row_count",
    "manual_review_row_count",
    "passed_row_count",
    "parse_error_row_count",
    "report_item_count",
    "verified_citation_count",
    "unavailable_citation_count",
    "high_attention_row_count",
    "manual_attention_row_count",
    "cleared_row_count",
)

ITEM_BASE_HEADERS: Final = (
    "attention_group",
    "source_verdict",
    "source_outcome",
    "row_no",
    "finding_id",
    "rule_id",
    "rule_version",
    "reason_code",
    "reasoning",
    "citation_status",
    "requires_manual_citation",
    "source_content_sha256",
)
CITATION_FIELDS: Final = (
    "binding_id",
    "family_stable_key",
    "document_title",
    "document_version",
    "effective_date",
    "expiry_date",
    "clause_id",
    "clause_no",
    "hierarchy",
    "quote",
    "quote_start",
    "quote_end",
    "quote_sha256",
)
ITEM_HEADERS: Final = ITEM_BASE_HEADERS + tuple(
    f"citation_{order}_{field}" for order in range(1, 4) for field in CITATION_FIELDS
)
RAW_PREFIX_HEADERS: Final = ("file_version_id", "source_content_sha256", "row_no")
PARSE_ERROR_HEADERS: Final = (
    "row_no",
    "error_code",
    "column_name",
    "message",
    "source_content_sha256",
)
POLICY_HEADERS: Final = (
    "report_item_id",
    "citation_order",
    "binding_id",
    "family_id",
    "family_stable_key",
    "document_id",
    "document_title",
    "document_version",
    "effective_date",
    "expiry_date",
    "document_content_sha256",
    "clause_id",
    "clause_no",
    "hierarchy",
    "clause_text",
    "clause_text_sha256",
    "quote",
    "quote_start",
    "quote_end",
    "quote_sha256",
    "verification_status",
)


class ReportXlsxError(ExpenseGuardError):
    """Stable failure raised before an XLSX artifact may be published."""

    status_code = 409


@dataclass(frozen=True)
class ReportFileMetadata:
    filename: str
    revision_no: int


@dataclass(frozen=True)
class RawEvidenceRow:
    row_no: int
    raw_json: Mapping[str, Any]


@dataclass(frozen=True)
class ReportWorkbookData:
    snapshot: ReportSnapshot
    file: ReportFileMetadata
    source_headers: tuple[str, ...]
    raw_rows: tuple[RawEvidenceRow, ...]
    generated_at: datetime


@dataclass(frozen=True)
class WorkbookExpectation:
    headers: Mapping[str, tuple[str, ...]]
    row_counts: Mapping[str, int]


def build_report_xlsx(data: ReportWorkbookData) -> bytes:
    """Build and mechanically verify the five-sheet F4 XLSX artifact."""
    _validate_raw_evidence(data)
    workbook = Workbook()
    summary_sheet = cast("Worksheet", workbook.active)
    summary_sheet.title = SHEET_NAMES[0]
    item_sheet = workbook.create_sheet(SHEET_NAMES[1])
    raw_sheet = workbook.create_sheet(SHEET_NAMES[2])
    error_sheet = workbook.create_sheet(SHEET_NAMES[3])
    policy_sheet = workbook.create_sheet(SHEET_NAMES[4])

    _write_table(summary_sheet, SUMMARY_HEADERS, _summary_rows(data))
    _write_table(item_sheet, ITEM_HEADERS, _item_rows(data.snapshot))
    raw_headers = RAW_PREFIX_HEADERS + data.source_headers
    _write_table(raw_sheet, raw_headers, _raw_rows(data))
    _write_table(error_sheet, PARSE_ERROR_HEADERS, _parse_error_rows(data.snapshot))
    _write_table(policy_sheet, POLICY_HEADERS, _policy_rows(data.snapshot))

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    verify_report_xlsx(content, workbook_expectation(data))
    return content


def verify_report_xlsx(content: bytes, expectation: WorkbookExpectation) -> None:
    """Fail closed unless the artifact exactly matches the safe workbook contract."""
    try:
        with ZipFile(BytesIO(content)) as archive:
            names = tuple(name.lower() for name in archive.namelist())
            forbidden_names = ("vbaproject", "externallinks/", "embeddings/", "oleobject")
            if any(token in name for name in names for token in forbidden_names):
                raise ReportXlsxError(
                    code="REPORT_EXPORT_UNSAFE_CONTENT", message="导出文件包含禁止的嵌入内容"
                )
            for name in archive.namelist():
                if not name.lower().endswith((".xml", ".rels")):
                    continue
                payload = archive.read(name).lower()
                if any(
                    marker in payload
                    for marker in (b"ddelink", b"externallink", b"oleobject", b"hyperlink")
                ):
                    raise ReportXlsxError(
                        code="REPORT_EXPORT_UNSAFE_CONTENT", message="导出文件包含禁止的链接内容"
                    )
    except BadZipFile as exc:
        raise ReportXlsxError(
            code="REPORT_EXPORT_INVALID_ARTIFACT", message="导出文件不是有效的 XLSX"
        ) from exc

    try:
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    except (BadZipFile, OSError, ValueError) as exc:
        raise ReportXlsxError(
            code="REPORT_EXPORT_INVALID_ARTIFACT", message="导出文件无法回读"
        ) from exc
    try:
        if tuple(workbook.sheetnames) != SHEET_NAMES:
            raise ReportXlsxError(
                code="REPORT_EXPORT_VERIFICATION_FAILED", message="导出工作表结构校验失败"
            )
        for sheet_name in SHEET_NAMES:
            sheet = workbook[sheet_name]
            actual_headers = tuple(
                str(cell.value) if cell.value is not None else "" for cell in sheet[1]
            )
            if actual_headers != expectation.headers[sheet_name]:
                raise ReportXlsxError(
                    code="REPORT_EXPORT_VERIFICATION_FAILED", message="导出标题列校验失败"
                )
            if sheet.max_row - 1 != expectation.row_counts[sheet_name]:
                raise ReportXlsxError(
                    code="REPORT_EXPORT_VERIFICATION_FAILED", message="导出行数校验失败"
                )
            if sheet.freeze_panes != "A2" or sheet.auto_filter.ref is None:
                raise ReportXlsxError(
                    code="REPORT_EXPORT_VERIFICATION_FAILED", message="导出可读性设置校验失败"
                )
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or cell.hyperlink is not None:
                        raise ReportXlsxError(
                            code="REPORT_EXPORT_UNSAFE_CONTENT",
                            message="导出文件包含禁止的公式或链接",
                        )
                    if isinstance(cell.value, str) and len(cell.value) > MAX_CELL_CHARACTERS:
                        raise ReportXlsxError(
                            code="REPORT_EXPORT_CELL_TOO_LONG", message="导出单元格超过长度上限"
                        )
    finally:
        workbook.close()


def safe_xlsx_value(value: object) -> CellValue:
    """Encode a cell without mutating the persisted source value."""
    if value is None or isinstance(value, bool | int | float | date | datetime):
        return value
    if isinstance(value, (dict, list, tuple)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    else:
        text = str(value)
    if len(text) > MAX_CELL_CHARACTERS:
        raise ReportXlsxError(code="REPORT_EXPORT_CELL_TOO_LONG", message="导出单元格超过长度上限")
    inspected = text.lstrip(" \t\r\n")
    if inspected.startswith(("=", "+", "-", "@")):
        if len(text) == MAX_CELL_CHARACTERS:
            raise ReportXlsxError(
                code="REPORT_EXPORT_CELL_TOO_LONG", message="公式防护后单元格超过长度上限"
            )
        return f"'{text}"
    return text


def workbook_expectation(data: ReportWorkbookData) -> WorkbookExpectation:
    """Return the exact structural contract for an encoded report."""
    return WorkbookExpectation(
        headers={
            SHEET_NAMES[0]: SUMMARY_HEADERS,
            SHEET_NAMES[1]: ITEM_HEADERS,
            SHEET_NAMES[2]: RAW_PREFIX_HEADERS + data.source_headers,
            SHEET_NAMES[3]: PARSE_ERROR_HEADERS,
            SHEET_NAMES[4]: POLICY_HEADERS,
        },
        row_counts={
            SHEET_NAMES[0]: len(SUMMARY_METRICS),
            SHEET_NAMES[1]: len(data.snapshot.items),
            SHEET_NAMES[2]: len(data.raw_rows),
            SHEET_NAMES[3]: len(data.snapshot.parse_errors),
            SHEET_NAMES[4]: sum(len(item.citations) for item in data.snapshot.items),
        },
    )


def _write_table(
    sheet: Worksheet, headers: Sequence[str], rows: Iterable[Sequence[object]]
) -> None:
    sheet.append(tuple(headers))
    for row in rows:
        sheet.append(tuple(safe_xlsx_value(value) for value in row))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{_column_letter(len(headers))}{max(sheet.max_row, 1)}"
    for row in sheet.iter_rows():
        for cell in row:
            _apply_number_format(cast("Cell", cell))


def _summary_rows(data: ReportWorkbookData) -> tuple[tuple[object, object], ...]:
    summary = data.snapshot.summary
    generated_at = _excel_datetime(data.generated_at)
    values: dict[str, object] = {
        "file_version_id": str(summary.file_version_id),
        "filename": data.file.filename,
        "revision_no": data.file.revision_no,
        "source_content_sha256": summary.source_content_sha256,
        "validation_run_id": str(summary.validation_run_id),
        "report_run_id": str(summary.report_run_id),
        "report_fingerprint": summary.report_fingerprint,
        "ruleset_fingerprint": summary.ruleset_fingerprint,
        "mapping_version_id": str(summary.mapping_version_id),
        "policy_fingerprint": canonical_sha256(data.snapshot.policy_manifest),
        "binding_fingerprint": canonical_sha256(data.snapshot.binding_manifest),
        "report_template_version": summary.template_version,
        "export_template_version": XLSX_TEMPLATE_VERSION,
        "generated_at": generated_at,
        "stored_row_count": summary.stored_row_count,
        "validated_row_count": summary.validated_row_count,
        "flagged_row_count": summary.flagged_row_count,
        "manual_review_row_count": summary.manual_review_row_count,
        "passed_row_count": summary.passed_row_count,
        "parse_error_row_count": summary.parse_error_row_count,
        "report_item_count": summary.report_item_count,
        "verified_citation_count": summary.verified_citation_count,
        "unavailable_citation_count": summary.unavailable_citation_count,
        "high_attention_row_count": summary.high_attention_row_count,
        "manual_attention_row_count": summary.manual_attention_row_count,
        "cleared_row_count": summary.cleared_row_count,
    }
    return tuple((metric, values[metric]) for metric in SUMMARY_METRICS)


def _item_rows(snapshot: ReportSnapshot) -> tuple[tuple[object, ...], ...]:
    result: list[tuple[object, ...]] = []
    for item in snapshot.items:
        values: list[object] = [
            item.attention_group.value,
            item.source_verdict,
            item.source_outcome,
            item.row_no,
            str(item.finding_id),
            item.rule_id,
            item.rule_version,
            item.reason_code,
            item.reasoning_snapshot,
            item.citation_status.value,
            item.requires_manual_citation,
            item.source_content_sha256,
        ]
        citations = {citation.citation_order: citation for citation in item.citations}
        for order in range(1, 4):
            citation = citations.get(order)
            if citation is None:
                values.extend([None] * len(CITATION_FIELDS))
            else:
                values.extend(
                    [
                        str(citation.binding_id),
                        citation.family_stable_key,
                        citation.document_title,
                        citation.document_version,
                        citation.effective_date,
                        citation.expiry_date,
                        str(citation.policy_clause_id),
                        citation.clause_no,
                        citation.hierarchy_path,
                        citation.quote,
                        citation.quote_start,
                        citation.quote_end,
                        citation.quote_sha256,
                    ]
                )
        result.append(tuple(values))
    return tuple(result)


def _raw_rows(data: ReportWorkbookData) -> tuple[tuple[object, ...], ...]:
    summary = data.snapshot.summary
    return tuple(
        (
            str(summary.file_version_id),
            summary.source_content_sha256,
            row.row_no,
            *(row.raw_json[header] for header in data.source_headers),
        )
        for row in data.raw_rows
    )


def _parse_error_rows(snapshot: ReportSnapshot) -> tuple[tuple[object, ...], ...]:
    return tuple(
        (
            error.row_no,
            error.error_code,
            error.column_name,
            error.message,
            error.source_content_sha256,
        )
        for error in snapshot.parse_errors
    )


def _policy_rows(snapshot: ReportSnapshot) -> tuple[tuple[object, ...], ...]:
    return tuple(
        (
            str(item.id),
            citation.citation_order,
            str(citation.binding_id),
            str(citation.policy_family_id),
            citation.family_stable_key,
            str(citation.policy_document_id),
            citation.document_title,
            citation.document_version,
            citation.effective_date,
            citation.expiry_date,
            citation.document_content_sha256,
            str(citation.policy_clause_id),
            citation.clause_no,
            citation.hierarchy_path,
            citation.clause_text,
            citation.clause_text_sha256,
            citation.quote,
            citation.quote_start,
            citation.quote_end,
            citation.quote_sha256,
            "verified_exact",
        )
        for item in snapshot.items
        for citation in item.citations
    )


def _validate_raw_evidence(data: ReportWorkbookData) -> None:
    if not data.source_headers or len(set(data.source_headers)) != len(data.source_headers):
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_HEADER_INVALID", message="原始文件标题结构不可用"
        )
    expected_rows = sorted(
        {item.row_no for item in data.snapshot.items}
        | {error.row_no for error in data.snapshot.parse_errors}
    )
    actual_rows = [row.row_no for row in data.raw_rows]
    if actual_rows != expected_rows:
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_EVIDENCE_INVALID", message="原始行证据不完整"
        )
    expected_headers = set(data.source_headers)
    if any(set(row.raw_json) != expected_headers for row in data.raw_rows):
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_HEADER_INVALID", message="原始行标题与源文件不一致"
        )


def _excel_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(UTC).replace(tzinfo=None)


def _apply_number_format(cell: Cell) -> None:
    if isinstance(cell.value, datetime):
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
    elif isinstance(cell.value, date):
        cell.number_format = "yyyy-mm-dd"


def _column_letter(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result
