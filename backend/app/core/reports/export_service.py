"""Crash-recoverable XLSX artifact generation and audited downloads."""

from __future__ import annotations

import hashlib
import os
import uuid
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Final, cast
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.exc import OperationalError
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.errors import ExpenseGuardError, NotFoundError
from app.core.policies.canonical import canonical_sha256
from app.core.reports.service import idempotency_key_hash, load_report_snapshot
from app.core.reports.xlsx import (
    XLSX_TEMPLATE_VERSION,
    RawEvidenceRow,
    ReportFileMetadata,
    ReportWorkbookData,
    ReportXlsxError,
    build_report_xlsx,
    verify_report_xlsx,
    workbook_expectation,
)
from app.core.security.auth_service import write_audit
from app.core.tenancy.locking import lock_tenant_nowait
from app.core.tenancy.scope import bind_tenant
from app.db.base import utc_now
from app.db.models.batch import ExpenseRow, FileVersion
from app.db.models.reports import ReportExport, ReportExportStatus, ReportRun, ReportRunStatus
from app.db.models.tenancy import AppUser

EXPORT_FORMAT: Final = "xlsx"


class ReportExportError(ExpenseGuardError):
    """Stable export-domain failure."""

    status_code = 409


class ReportExportInternalError(ReportExportError):
    status_code = 500

    def __init__(self) -> None:
        super().__init__(code="REPORT_EXPORT_INTERNAL_ERROR", message="报告导出失败，请稍后重试")


@dataclass(frozen=True)
class ReportExportSummary:
    export_id: uuid.UUID
    report_run_id: uuid.UUID
    format: str
    template_version: str
    artifact_sha256: str
    size_bytes: int
    completed_at: datetime
    reused_existing: bool


@dataclass(frozen=True)
class DownloadedReportArtifact:
    export_id: uuid.UUID
    filename: str
    content_type: str
    artifact_sha256: str
    content: bytes


async def create_report_export(
    db: AsyncSession,
    session_factory: async_sessionmaker[AsyncSession],
    *,
    tenant_id: uuid.UUID,
    actor_id: uuid.UUID,
    report_run_id: uuid.UUID,
    idempotency_key: str,
    export_root: Path,
    upload_root: Path,
    template_version: str = XLSX_TEMPLATE_VERSION,
) -> ReportExportSummary:
    """Create or replay the sole XLSX artifact for a completed report."""
    key_hash = idempotency_key_hash(idempotency_key)
    request_fingerprint = _request_fingerprint(report_run_id, template_version)
    export_id: uuid.UUID | None = None
    try:
        await lock_tenant_nowait(db, tenant_id)
        actor = await db.scalar(select(AppUser).where(AppUser.id == actor_id))
        if actor is None:
            raise NotFoundError(code="REPORT_EXPORT_ACTOR_NOT_FOUND", message="操作人不存在")
        report = await db.scalar(
            select(ReportRun).where(
                ReportRun.id == report_run_id, ReportRun.status == ReportRunStatus.COMPLETED
            )
        )
        if report is None:
            raise NotFoundError(code="REPORT_NOT_FOUND", message="报告不存在")

        keyed = await db.scalar(
            select(ReportExport).where(ReportExport.idempotency_key_hash == key_hash)
        )
        if keyed is not None and keyed.request_fingerprint != request_fingerprint:
            raise ReportExportError(
                code="IDEMPOTENCY_KEY_REUSED", message="该 Idempotency-Key 已绑定其他导出请求"
            )
        export = keyed or await db.scalar(
            select(ReportExport).where(
                ReportExport.report_run_id == report_run_id,
                ReportExport.format == EXPORT_FORMAT,
                ReportExport.template_version == template_version,
            )
        )
        if export is not None and export.status is ReportExportStatus.COMPLETED:
            content = _read_completed_artifact(export, export_root)
            return _summary(export, content=content, reused_existing=True)
        if export is None:
            export = ReportExport(
                tenant_id=tenant_id,
                report_run_id=report_run_id,
                format=EXPORT_FORMAT,
                template_version=template_version,
                status=ReportExportStatus.IN_PROGRESS,
                request_fingerprint=request_fingerprint,
                idempotency_key_hash=key_hash,
                created_by=actor_id,
            )
            db.add(export)
            await db.flush()
        else:
            if export.request_fingerprint != request_fingerprint:
                raise ReportExportError(
                    code="REPORT_EXPORT_REQUEST_CONFLICT", message="该报告已绑定其他导出请求"
                )
            export.status = ReportExportStatus.IN_PROGRESS
            export.artifact_storage_key = None
            export.artifact_sha256 = None
            export.size_bytes = None
            export.completed_at = None
            export.failure_code = None
            await db.flush()
        export_id = export.id

        storage_key = _artifact_storage_key(tenant_id, report_run_id, template_version)
        target = _resolve_storage_key(export_root, storage_key)
        workbook_data = await _load_export_data(db=db, report=report, upload_root=upload_root)
        content = _recover_or_build(target=target, data=workbook_data)
        digest = hashlib.sha256(content).hexdigest()
        export.status = ReportExportStatus.COMPLETED
        export.artifact_storage_key = storage_key
        export.artifact_sha256 = digest
        export.size_bytes = len(content)
        export.completed_at = utc_now()
        await write_audit(
            db,
            tenant_id=tenant_id,
            actor_id=actor_id,
            action="report.export_generate",
            target_type="report_export",
            target_id=str(export.id),
            payload={
                "report_run_id": str(report_run_id),
                "artifact_sha256": digest,
                "template_version": template_version,
                "size_bytes": len(content),
            },
        )
        await db.flush()
        return _summary(export, content=content, reused_existing=False)
    except (NotFoundError, ReportExportError):
        raise
    except OperationalError as exc:
        if _sqlstate(exc) == "55P03":
            raise ReportExportError(
                code="REPORT_EXPORT_IN_PROGRESS",
                message="该租户已有导出正在生成，请稍后重试",
            ) from exc
        await _rollback_and_record_failure(
            db,
            session_factory,
            tenant_id=tenant_id,
            actor_id=actor_id,
            report_run_id=report_run_id,
            export_id=export_id,
            key_hash=key_hash,
            request_fingerprint=request_fingerprint,
            template_version=template_version,
            failure_code="REPORT_EXPORT_INTERNAL_ERROR",
        )
        raise ReportExportInternalError from exc
    except ReportXlsxError as exc:
        await _rollback_and_record_failure(
            db,
            session_factory,
            tenant_id=tenant_id,
            actor_id=actor_id,
            report_run_id=report_run_id,
            export_id=export_id,
            key_hash=key_hash,
            request_fingerprint=request_fingerprint,
            template_version=template_version,
            failure_code=exc.code,
        )
        raise ReportExportError(code=exc.code, message=exc.message) from exc
    except Exception as exc:
        await _rollback_and_record_failure(
            db,
            session_factory,
            tenant_id=tenant_id,
            actor_id=actor_id,
            report_run_id=report_run_id,
            export_id=export_id,
            key_hash=key_hash,
            request_fingerprint=request_fingerprint,
            template_version=template_version,
            failure_code="REPORT_EXPORT_INTERNAL_ERROR",
        )
        raise ReportExportInternalError from exc


async def download_report_export(
    db: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    actor_id: uuid.UUID,
    export_id: uuid.UUID,
    export_root: Path,
) -> DownloadedReportArtifact:
    """Read, hash-check and audit one authorized completed artifact download."""
    actor = await db.scalar(select(AppUser).where(AppUser.id == actor_id))
    if actor is None:
        raise NotFoundError(code="REPORT_EXPORT_ACTOR_NOT_FOUND", message="操作人不存在")
    export = await db.get(ReportExport, export_id)
    if export is None or export.status is not ReportExportStatus.COMPLETED:
        raise NotFoundError(code="REPORT_EXPORT_NOT_FOUND", message="导出文件不存在")
    content = _read_completed_artifact(export, export_root)
    report = await db.get(ReportRun, export.report_run_id)
    if report is None:
        raise NotFoundError(code="REPORT_NOT_FOUND", message="报告不存在")
    batch = await db.get(FileVersion, report.file_version_id)
    if batch is None:
        raise NotFoundError(code="BATCH_NOT_FOUND", message="批次不存在")
    await write_audit(
        db,
        tenant_id=tenant_id,
        actor_id=actor_id,
        action="report.export_download",
        target_type="report_export",
        target_id=str(export.id),
        payload={
            "report_run_id": str(report.id),
            "artifact_sha256": cast("str", export.artifact_sha256),
            "size_bytes": len(content),
        },
    )
    await db.flush()
    return DownloadedReportArtifact(
        export_id=export.id,
        filename=_download_filename(batch.filename),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        artifact_sha256=cast("str", export.artifact_sha256),
        content=content,
    )


async def _load_export_data(
    *, db: AsyncSession, report: ReportRun, upload_root: Path
) -> ReportWorkbookData:
    snapshot = await load_report_snapshot(db, report_run_id=report.id)
    batch = await db.get(FileVersion, report.file_version_id)
    if batch is None:
        raise ReportXlsxError(code="REPORT_EXPORT_SOURCE_UNAVAILABLE", message="原始批次不可读取")
    source_headers = _load_source_headers(
        upload_root=upload_root,
        tenant_id=report.tenant_id,
        content_sha256=report.source_content_sha256,
    )
    referenced = sorted(
        {item.row_no for item in snapshot.items} | {error.row_no for error in snapshot.parse_errors}
    )
    rows: tuple[ExpenseRow, ...]
    if referenced:
        rows = tuple(
            (
                await db.scalars(
                    select(ExpenseRow)
                    .where(
                        ExpenseRow.file_version_id == report.file_version_id,
                        ExpenseRow.row_no.in_(referenced),
                    )
                    .order_by(ExpenseRow.row_no)
                )
            ).all()
        )
    else:
        rows = ()
    return ReportWorkbookData(
        snapshot=snapshot,
        file=ReportFileMetadata(filename=batch.filename, revision_no=batch.revision_no),
        source_headers=source_headers,
        raw_rows=tuple(RawEvidenceRow(row_no=row.row_no, raw_json=row.raw_json) for row in rows),
        # The export's semantic generation timestamp is anchored to the immutable
        # report completion. That keeps crash recovery comparable without claiming
        # byte-identical ZIP packaging across library versions.
        generated_at=cast("datetime", report.completed_at),
    )


def _load_source_headers(
    *, upload_root: Path, tenant_id: uuid.UUID, content_sha256: str
) -> tuple[str, ...]:
    source = (upload_root.resolve() / str(tenant_id) / f"{content_sha256}.xlsx").resolve()
    if not source.is_relative_to(upload_root.resolve()):
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_UNAVAILABLE", message="原始批次存储路径无效"
        )
    try:
        content = source.read_bytes()
    except OSError as exc:
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_UNAVAILABLE", message="原始批次文件不可读取"
        ) from exc
    if hashlib.sha256(content).hexdigest() != content_sha256:
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_HASH_MISMATCH", message="原始批次文件校验失败"
        )
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            first = next(
                workbook.worksheets[0].iter_rows(min_row=1, max_row=1, values_only=True), None
            )
        finally:
            workbook.close()
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_UNAVAILABLE", message="原始批次标题不可读取"
        ) from exc
    if first is None:
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_HEADER_INVALID", message="原始批次缺少标题"
        )
    headers = tuple(str(value).strip() if value is not None else "" for value in first)
    if not headers or any(not value for value in headers) or len(set(headers)) != len(headers):
        raise ReportXlsxError(
            code="REPORT_EXPORT_SOURCE_HEADER_INVALID", message="原始批次标题结构无效"
        )
    return headers


def _recover_or_build(*, target: Path, data: ReportWorkbookData) -> bytes:
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        content = target.read_bytes()
        verify_report_xlsx(content, workbook_expectation(data))
        expected = build_report_xlsx(data)
        if _semantic_cells(content) != _semantic_cells(expected):
            raise ReportXlsxError(
                code="REPORT_EXPORT_ARTIFACT_CONFLICT",
                message="既有导出文件语义校验失败",
            )
        return content
    content = build_report_xlsx(data)
    temporary = target.with_name(f".{target.name}.{uuid.uuid4().hex}.tmp")
    try:
        temporary.write_bytes(content)
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)
    return content


def _read_completed_artifact(export: ReportExport, export_root: Path) -> bytes:
    if export.artifact_storage_key is None or export.artifact_sha256 is None:
        raise ReportExportError(
            code="REPORT_EXPORT_METADATA_INVALID", message="导出文件元数据不完整"
        )
    target = _resolve_storage_key(export_root, export.artifact_storage_key)
    try:
        content = target.read_bytes()
    except OSError as exc:
        raise ReportExportError(
            code="REPORT_EXPORT_ARTIFACT_MISSING", message="导出文件已丢失"
        ) from exc
    if hashlib.sha256(content).hexdigest() != export.artifact_sha256:
        raise ReportExportError(
            code="REPORT_EXPORT_ARTIFACT_HASH_MISMATCH", message="导出文件校验失败"
        )
    return content


def _semantic_cells(content: bytes) -> tuple[tuple[tuple[object, str, str], ...], ...]:
    """Compare workbook meaning while deliberately ignoring ZIP package bytes."""
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    try:
        return tuple(
            tuple(
                (cell.value, cell.data_type, cell.number_format)
                for row in workbook[sheet_name].iter_rows()
                for cell in row
            )
            for sheet_name in workbook.sheetnames
        )
    finally:
        workbook.close()


async def _rollback_and_record_failure(
    db: AsyncSession,
    session_factory: async_sessionmaker[AsyncSession],
    *,
    tenant_id: uuid.UUID,
    actor_id: uuid.UUID,
    report_run_id: uuid.UUID,
    export_id: uuid.UUID | None,
    key_hash: str,
    request_fingerprint: str,
    template_version: str,
    failure_code: str,
) -> None:
    await db.rollback()
    async with session_factory() as failure_db:
        bind_tenant(failure_db.sync_session, tenant_id)
        actor = await failure_db.scalar(select(AppUser).where(AppUser.id == actor_id))
        report = await failure_db.scalar(
            select(ReportRun).where(
                ReportRun.id == report_run_id, ReportRun.status == ReportRunStatus.COMPLETED
            )
        )
        if actor is None or report is None:
            return
        export = await failure_db.scalar(
            select(ReportExport).where(
                ReportExport.report_run_id == report_run_id,
                ReportExport.format == EXPORT_FORMAT,
                ReportExport.template_version == template_version,
            )
        )
        if export is None:
            export = ReportExport(
                id=export_id or uuid.uuid4(),
                tenant_id=tenant_id,
                report_run_id=report_run_id,
                format=EXPORT_FORMAT,
                template_version=template_version,
                status=ReportExportStatus.FAILED,
                request_fingerprint=request_fingerprint,
                idempotency_key_hash=key_hash,
                created_by=actor_id,
                completed_at=utc_now(),
                failure_code=failure_code,
            )
            failure_db.add(export)
        elif export.status is not ReportExportStatus.COMPLETED:
            export.status = ReportExportStatus.FAILED
            export.artifact_storage_key = None
            export.artifact_sha256 = None
            export.size_bytes = None
            export.completed_at = utc_now()
            export.failure_code = failure_code
        await write_audit(
            failure_db,
            tenant_id=tenant_id,
            actor_id=actor_id,
            action="report.export_failed",
            target_type="report_export",
            target_id=str(export.id),
            payload={"report_run_id": str(report_run_id), "failure_code": failure_code},
        )
        await failure_db.commit()


def _summary(export: ReportExport, *, content: bytes, reused_existing: bool) -> ReportExportSummary:
    if export.artifact_sha256 is None or export.size_bytes is None or export.completed_at is None:
        raise RuntimeError("completed export metadata invalid")
    if len(content) != export.size_bytes:
        raise ReportExportError(
            code="REPORT_EXPORT_ARTIFACT_HASH_MISMATCH", message="导出文件大小校验失败"
        )
    return ReportExportSummary(
        export_id=export.id,
        report_run_id=export.report_run_id,
        format=export.format,
        template_version=export.template_version,
        artifact_sha256=export.artifact_sha256,
        size_bytes=export.size_bytes,
        completed_at=export.completed_at,
        reused_existing=reused_existing,
    )


def _request_fingerprint(report_run_id: uuid.UUID, template_version: str) -> str:
    return canonical_sha256(
        {
            "format": EXPORT_FORMAT,
            "report_run_id": str(report_run_id),
            "schema_version": 1,
            "template_version": template_version,
        }
    )


def _artifact_storage_key(
    tenant_id: uuid.UUID, report_run_id: uuid.UUID, template_version: str
) -> str:
    template_hash = hashlib.sha256(template_version.encode("utf-8")).hexdigest()[:16]
    return f"{tenant_id}/{report_run_id}/{template_hash}/report.xlsx"


def _resolve_storage_key(root: Path, storage_key: str) -> Path:
    resolved_root = root.expanduser().resolve()
    if not storage_key or Path(storage_key).is_absolute():
        raise ReportExportError(code="REPORT_EXPORT_STORAGE_KEY_INVALID", message="导出存储键无效")
    target = (resolved_root / storage_key).resolve()
    if not target.is_relative_to(resolved_root):
        raise ReportExportError(code="REPORT_EXPORT_STORAGE_KEY_INVALID", message="导出存储键无效")
    return target


def _download_filename(source_filename: str) -> str:
    stem = Path(source_filename).stem
    safe = "".join(
        character for character in stem if character >= " " and character not in '\\/:*?"<>|'
    )
    safe = safe.strip(" .") or "批次"
    return f"费用预审报告-{safe}.xlsx"


def _sqlstate(exc: OperationalError) -> str | None:
    value: object = getattr(exc.orig, "sqlstate", None)
    return value if isinstance(value, str) else None
