"""Excel 导入与文件版本管理服务。"""

from __future__ import annotations

import hashlib
import os
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, cast
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.errors import ExpenseGuardError, NotFoundError
from app.db.models.batch import ExpenseRow, FileVersion

MIN_IMPORT_ROWS = 500
MAX_IMPORT_ROWS = 5000


class BatchImportError(ExpenseGuardError):
    """批次导入输入不合法。"""


@dataclass(frozen=True)
class ParsedExpenseRow:
    """Excel 中一条可落库的原始行。"""

    row_no: int
    raw_json: dict[str, Any]


@dataclass(frozen=True)
class ParsedWorkbook:
    """已解析但尚未落库的 workbook。"""

    rows: tuple[ParsedExpenseRow, ...]

    @property
    def row_count(self) -> int:
        """数据行数量,不含表头。"""
        return len(self.rows)


@dataclass(frozen=True)
class BatchSummary:
    """批次元数据。"""

    file_version_id: uuid.UUID
    filename: str
    content_hash: str
    row_count: int
    uploaded_at: datetime
    uploaded_by: uuid.UUID


@dataclass(frozen=True)
class BatchImportResult:
    """导入响应。"""

    summary: BatchSummary
    reused_existing: bool
    stored_rows: int


@dataclass(frozen=True)
class ExpenseRowSummary:
    """批次详情中的原始行摘要。"""

    row_no: int
    raw_json: dict[str, Any]
    parse_error: str | None


@dataclass(frozen=True)
class BatchDetail:
    """批次详情。"""

    summary: BatchSummary
    rows: tuple[ExpenseRowSummary, ...]


def sha256_hex(content: bytes) -> str:
    """计算文件内容 SHA-256。"""
    return hashlib.sha256(content).hexdigest()


def parse_xlsx(content: bytes) -> ParsedWorkbook:
    """解析 `.xlsx` 为原始 JSON 行。

    F1 不做业务字段归一化,只保留证据链。第一行是表头;表头后的非全空
    物理行进入 `expense_row`,并保留 Excel 行号。
    """
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise BatchImportError(
            code="BATCH_XLSX_INVALID",
            message="无法读取该 Excel 文件,请确认文件为有效的 .xlsx",
        ) from exc

    try:
        sheet = workbook.worksheets[0]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = _parse_headers(header_cells)
        rows: list[ParsedExpenseRow] = []
        data_rows = sheet.iter_rows(min_row=2, values_only=True)
        for excel_row_no, values in enumerate(data_rows, start=2):
            if values is None or all(_is_blank(value) for value in values):
                continue
            raw = {
                header: cell_to_json(values[index] if index < len(values) else None)
                for index, header in enumerate(headers)
            }
            rows.append(ParsedExpenseRow(row_no=excel_row_no, raw_json=raw))
    finally:
        workbook.close()

    if len(rows) < MIN_IMPORT_ROWS:
        raise BatchImportError(
            code="BATCH_ROW_COUNT_TOO_LOW",
            message=f"Excel 至少需要 {MIN_IMPORT_ROWS} 行数据",
        )
    if len(rows) > MAX_IMPORT_ROWS:
        raise BatchImportError(
            code="BATCH_ROW_COUNT_TOO_HIGH",
            message=f"Excel 最多支持 {MAX_IMPORT_ROWS} 行数据",
        )
    return ParsedWorkbook(rows=tuple(rows))


def cell_to_json(value: object) -> str | int | float | bool | None:
    """把 openpyxl 单元格值转换为 JSONB 可安全存储的值。"""
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date | time):
        return value.isoformat()
    return str(value)


async def import_batch(
    db: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    uploaded_by: uuid.UUID,
    filename: str,
    content: bytes,
    upload_root: Path,
) -> BatchImportResult:
    """导入 Excel 文件;同租户同内容复用既有 file_version。"""
    _validate_filename(filename)
    content_hash = sha256_hex(content)

    existing = await _get_existing(db, content_hash)
    if existing is not None:
        return BatchImportResult(
            summary=_to_summary(existing),
            reused_existing=True,
            stored_rows=await _count_rows(db, existing.id),
        )

    parsed = parse_xlsx(content)
    _store_original_file(
        upload_root=upload_root,
        tenant_id=tenant_id,
        content_hash=content_hash,
        content=content,
    )

    try:
        file_version = FileVersion(
            tenant_id=tenant_id,
            filename=filename,
            content_hash=content_hash,
            row_count=parsed.row_count,
            uploaded_by=uploaded_by,
        )
        db.add(file_version)
        await db.flush()

        db.add_all(
            ExpenseRow(
                tenant_id=tenant_id,
                file_version_id=file_version.id,
                row_no=row.row_no,
                raw_json=row.raw_json,
                parse_error=None,
            )
            for row in parsed.rows
        )
        await db.flush()
    except IntegrityError:
        await db.rollback()
        existing_after_conflict = await _get_existing(db, content_hash)
        if existing_after_conflict is None:
            raise
        return BatchImportResult(
            summary=_to_summary(existing_after_conflict),
            reused_existing=True,
            stored_rows=await _count_rows(db, existing_after_conflict.id),
        )

    await db.refresh(file_version)
    return BatchImportResult(
        summary=_to_summary(file_version),
        reused_existing=False,
        stored_rows=parsed.row_count,
    )


async def list_batches(db: AsyncSession) -> tuple[BatchSummary, ...]:
    """列出当前租户可见批次。"""
    rows = (await db.scalars(select(FileVersion).order_by(FileVersion.uploaded_at.desc()))).all()
    return tuple(_to_summary(row) for row in rows)


async def get_batch_detail(db: AsyncSession, file_version_id: uuid.UUID) -> BatchDetail:
    """读取批次详情与原始行。"""
    file_version = await db.get(FileVersion, file_version_id)
    if file_version is None:
        raise NotFoundError(code="BATCH_NOT_FOUND", message="批次不存在")
    rows = (
        await db.scalars(
            select(ExpenseRow)
            .where(ExpenseRow.file_version_id == file_version_id)
            .order_by(ExpenseRow.row_no)
        )
    ).all()
    return BatchDetail(
        summary=_to_summary(file_version),
        rows=tuple(
            ExpenseRowSummary(
                row_no=row.row_no,
                raw_json=row.raw_json,
                parse_error=row.parse_error,
            )
            for row in rows
        ),
    )


def _parse_headers(header_cells: tuple[object, ...] | None) -> tuple[str, ...]:
    if header_cells is None:
        raise BatchImportError(code="BATCH_HEADER_MISSING", message="Excel 缺少表头行")

    headers: list[str] = []
    seen: set[str] = set()
    for cell in header_cells:
        if cell is None or str(cell).strip() == "":
            raise BatchImportError(
                code="BATCH_HEADER_EMPTY",
                message="Excel 表头不能包含空列名",
            )
        header = str(cell).strip()
        if header in seen:
            raise BatchImportError(
                code="BATCH_HEADER_DUPLICATED",
                message=f"Excel 表头存在重复列名: {header}",
            )
        seen.add(header)
        headers.append(header)

    if not headers:
        raise BatchImportError(code="BATCH_HEADER_MISSING", message="Excel 缺少表头行")
    return tuple(headers)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _validate_filename(filename: str) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise BatchImportError(
            code="BATCH_FILE_TYPE_UNSUPPORTED",
            message="仅支持 .xlsx 文件",
        )


def _store_original_file(
    *,
    upload_root: Path,
    tenant_id: uuid.UUID,
    content_hash: str,
    content: bytes,
) -> None:
    target_dir = upload_root / str(tenant_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{content_hash}.xlsx"
    if target.exists():
        return
    with NamedTemporaryFile(delete=False, dir=target_dir, suffix=".tmp") as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)
    try:
        os.replace(tmp_path, target)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


async def _get_existing(db: AsyncSession, content_hash: str) -> FileVersion | None:
    return cast(
        "FileVersion | None",
        await db.scalar(
            select(FileVersion).where(
                FileVersion.content_hash == content_hash,
                FileVersion.revision_no == 1,
            )
        ),
    )


async def _count_rows(db: AsyncSession, file_version_id: uuid.UUID) -> int:
    count = await db.scalar(
        select(func.count())
        .select_from(ExpenseRow)
        .where(ExpenseRow.file_version_id == file_version_id)
    )
    return int(count or 0)


def _to_summary(file_version: FileVersion) -> BatchSummary:
    return BatchSummary(
        file_version_id=file_version.id,
        filename=file_version.filename,
        content_hash=file_version.content_hash,
        row_count=file_version.row_count or 0,
        uploaded_at=file_version.uploaded_at,
        uploaded_by=file_version.uploaded_by,
    )
