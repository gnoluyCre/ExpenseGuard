"""把合成批次落盘成三个**物理分离**的文件。

    <stem>.xlsx           只有数据，没有任何标签列  ← 喂给被测系统的就是它
    <stem>.labels.jsonl   每行一条真值标签
    <stem>.manifest.json  seed 与生成参数，用于重放

分成三个文件不是整理癖。标签泄漏的典型形态是「标签列悄悄留在了同一张表
里」，规则或模型于是间接看到了答案，评测指标全面虚高且很难察觉。物理分离
让泄漏需要一次显式的读文件动作才可能发生。

`_assert_no_label_leak` 在写盘前机械地再验一次:数据表的列必须**恰好**是
`DATA_COLUMNS`。多一列少一列都当场失败，而不是等评测结果好得可疑时才回头查。
"""

from __future__ import annotations

import json
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.synth.generator import DATA_COLUMNS, SyntheticBatch

#: 写进 xlsx 的日期格式。与 F2 的日期归一化对齐前先固定一种，避免夹具本身
#: 就带上「多种日期格式」这个变量——那属于 F2 要专门造的用例，不该混进基线。
_DATE_FORMAT = "yyyy-mm-dd"


@dataclass(frozen=True, slots=True)
class BatchFiles:
    """落盘产物的路径。"""

    data: Path
    labels: Path
    manifest: Path


def _assert_no_label_leak(rows: list[dict[str, Any]]) -> None:
    """确认数据行里没有混进标签字段。"""
    expected = set(DATA_COLUMNS)
    for index, row in enumerate(rows, start=1):
        actual = set(row)
        if actual != expected:
            extra = sorted(actual - expected)
            missing = sorted(expected - actual)
            raise ValueError(
                f"第 {index} 行的列与 DATA_COLUMNS 不一致 —— 多出 {extra}，缺少 {missing}。"
                "多出的列极可能是标签泄漏，不要通过放宽本检查来绕过。"
            )


def _cell_value(value: Any) -> Any:  # noqa: ANN401  (Excel 单元格本就是异构值)
    """转成 openpyxl 能写的类型。"""
    if isinstance(value, Decimal):
        # Excel 本身只有 IEEE754 浮点，没有十进制类型。这里的精度损失是
        # 文件格式固有的，不是本项目的取舍——真实客户表里的金额同样如此，
        # 所以夹具保持这个特性反而更贴近生产输入。
        return float(value)
    return value


def write_batch(batch: SyntheticBatch, out_dir: Path, stem: str = "batch") -> BatchFiles:
    """把一批数据写成 xlsx / jsonl / json 三个文件。"""
    _assert_no_label_leak(batch.rows)
    out_dir.mkdir(parents=True, exist_ok=True)

    files = BatchFiles(
        data=out_dir / f"{stem}.xlsx",
        labels=out_dir / f"{stem}.labels.jsonl",
        manifest=out_dir / f"{stem}.manifest.json",
    )

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover  (openpyxl 新建工作簿必有活动表)
        raise RuntimeError("openpyxl 未创建活动工作表")
    sheet.title = "报销明细"
    sheet.append(list(DATA_COLUMNS))
    for row in batch.rows:
        sheet.append([_cell_value(row[column]) for column in DATA_COLUMNS])
        for column_index, column in enumerate(DATA_COLUMNS, start=1):
            # datetime 是 date 的子类，判 date 即可覆盖两者
            if isinstance(row[column], date):
                sheet.cell(row=sheet.max_row, column=column_index).number_format = _DATE_FORMAT
    workbook.save(files.data)

    # jsonl 而非单个 json 数组:批次可能上万行，逐行读写不必把全部标签
    # 载入内存，且追加与 diff 都更友好。
    with files.labels.open("w", encoding="utf-8", newline="\n") as handle:
        for label in batch.labels:
            handle.write(label.model_dump_json() + "\n")

    files.manifest.write_text(
        json.dumps(
            batch.manifest.model_dump(mode="json"),
            indent=2,
            sort_keys=True,
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    return files


def read_labels(path: Path) -> Mapping[int, dict[str, Any]]:
    """读回标签，按 `row_no` 索引。评测侧的唯一入口。"""
    labels: dict[int, dict[str, Any]] = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip():
                continue
            record: dict[str, Any] = json.loads(line)
            labels[int(record["row_no"])] = record
    return labels
