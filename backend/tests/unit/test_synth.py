"""合成数据生成器的契约测试。

Phase 1 对生成器的要求是**确定性与诚实**，不是覆盖度。因此这里测的是:

- 同 seed 必然重放出同一批逻辑行，且**不受全局 `random` 污染**
- 登记表覆盖全部八类违规，未实现的那七类会当场炸而不是静默产出干净行
- 标签不出现在 `.xlsx` 里（标签泄漏的结构性预防）
"""

import json
import random
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.synth import (
    DATA_COLUMNS,
    GENERATOR_VERSION,
    IMPLEMENTED_KINDS,
    INJECTORS,
    ViolationKind,
    generate_batch,
    read_labels,
    write_batch,
)
from app.synth.kinds import CATEGORY_LIMITS

pytestmark = pytest.mark.unit

#: backend/tests/unit/test_synth.py → backend → 仓库根 → data/synthetic
_FIXTURE_DIR = Path(__file__).resolve().parents[3] / "data" / "synthetic"


def test_同种子重放出完全相同的逻辑行() -> None:
    first = generate_batch(seed=42, row_count=30)
    second = generate_batch(seed=42, row_count=30)

    assert first.rows == second.rows
    assert first.manifest.rows_digest == second.manifest.rows_digest
    assert [label.model_dump() for label in first.labels] == [
        label.model_dump() for label in second.labels
    ]


def test_不同种子产出不同数据() -> None:
    assert (
        generate_batch(seed=1, row_count=30).manifest.rows_digest
        != generate_batch(seed=2, row_count=30).manifest.rows_digest
    )


def test_全局random被污染时结果不变() -> None:
    """这条是「绝不用全局 `random`」的可执行证据。

    任何第三方库都可能调 `random.seed()`。若生成器用的是模块级全局实例，
    这里的两次结果就会不同——而那种 bug 表现为「昨天的夹具今天重放不出来」，
    极难归因。用 `Random(seed)` 显式实例则天然免疫。
    """
    baseline = generate_batch(seed=7, row_count=20).manifest.rows_digest

    random.seed(999)
    random.random()  # noqa: S311  (刻意污染全局状态，正是本测试要验的场景)
    polluted = generate_batch(seed=7, row_count=20).manifest.rows_digest

    assert polluted == baseline


def test_登记表覆盖全部违规类型() -> None:
    """无遗漏断言:新增 `ViolationKind` 成员却忘了登记，这里立刻红。"""
    assert set(INJECTORS) == set(ViolationKind)


def test_未实现的违规类型当场抛错() -> None:
    未实现 = sorted(set(ViolationKind) - IMPLEMENTED_KINDS)
    assert 未实现, "若八类都已实现，请删掉本测试并更新 IMPLEMENTED_KINDS 的文档"

    for kind in 未实现:
        with pytest.raises(NotImplementedError, match=kind.value):
            generate_batch(seed=1, row_count=5, violation_rate=1.0, kinds=(kind,))


def test_超限额行的金额确实超过该类型限额() -> None:
    batch = generate_batch(seed=11, row_count=40, violation_rate=1.0)

    assert all(label.is_violation for label in batch.labels)
    for row, label in zip(batch.rows, batch.labels, strict=True):
        detail = label.details[ViolationKind.OVER_LIMIT.value]
        limit = CATEGORY_LIMITS[row["费用类型"]]
        assert row["金额"] > limit
        # 标签里的数值必须与数据行一致——评测比对的是标签，二者漂移
        # 会制造静默的错误标注
        assert Decimal(detail["amount"]) == row["金额"]
        assert Decimal(detail["limit"]) == limit


def test_合规行金额不超限额() -> None:
    batch = generate_batch(seed=13, row_count=40, violation_rate=0.0)

    assert not any(label.is_violation for label in batch.labels)
    for row in batch.rows:
        assert row["金额"] <= CATEGORY_LIMITS[row["费用类型"]]


@pytest.mark.parametrize(
    ("kwargs", "reason"),
    [
        ({"row_count": 0}, "行数"),
        ({"violation_rate": 1.5}, "占比越界"),
        ({"kinds": ()}, "空类型"),
    ],
)
def test_非法参数直接报错(kwargs: dict[str, object], reason: str) -> None:
    with pytest.raises(ValueError):
        generate_batch(seed=1, **kwargs)  # type: ignore[arg-type]


def test_落盘后标签不出现在数据文件里(tmp_path: Path) -> None:
    batch = generate_batch(seed=5, row_count=25, violation_rate=0.4)
    files = write_batch(batch, tmp_path, stem="fixture")

    sheet = load_workbook(files.data).active
    assert sheet is not None
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # 恰好相等，不是「包含」——多一列就可能是泄漏
    assert header == list(DATA_COLUMNS)
    assert sheet.max_row == batch.manifest.row_count + 1  # 表头占一行

    # 数据文件里不得出现任何标签词汇
    all_text = " ".join(str(cell.value) for row in sheet.iter_rows() for cell in row)
    assert "is_violation" not in all_text
    assert ViolationKind.OVER_LIMIT.value not in all_text


def test_三个文件分离且可回读(tmp_path: Path) -> None:
    batch = generate_batch(seed=5, row_count=25, violation_rate=0.4)
    files = write_batch(batch, tmp_path, stem="fixture")

    assert files.data.exists() and files.labels.exists() and files.manifest.exists()

    labels = read_labels(files.labels)
    assert set(labels) == set(range(1, 26))

    manifest = json.loads(files.manifest.read_text(encoding="utf-8"))
    assert manifest["seed"] == 5
    assert manifest["rows_digest"] == batch.manifest.rows_digest
    # manifest 必须如实交代「还不会造哪几类」，否则下游会误以为覆盖齐全
    assert set(manifest["kinds_not_implemented"]) == {
        k.value for k in set(ViolationKind) - IMPLEMENTED_KINDS
    }


def test_已提交样本可由manifest重放() -> None:
    """提交进仓库的那份 50 行样本必须能用 manifest 里的 seed 原样重放。

    这条测试让「大批次不进仓库，用 seed 现场重放」这个提交策略成立:
    一旦有人改了生成器却没重新导出样本，重放结果与 `rows_digest` 不符，
    这里立刻红——而不是等某天有人拿旧夹具跑评测，得到一组对不上的数字。
    """
    manifest = json.loads((_FIXTURE_DIR / "baseline-50.manifest.json").read_text(encoding="utf-8"))

    assert manifest["generator_version"] == GENERATOR_VERSION, (
        "生成器契约版本已变更，样本必须重新导出:"
        "uv run python -m app.synth --seed "
        f"{manifest['seed']} --rows {manifest['row_count']} "
        "--out ../data/synthetic --stem baseline-50"
    )

    replayed = generate_batch(
        seed=manifest["seed"],
        row_count=manifest["row_count"],
        violation_rate=manifest["violation_rate"],
        kinds=tuple(ViolationKind(k) for k in manifest["kinds_requested"]),
        anchor_date=date.fromisoformat(manifest["anchor_date"]),
    )
    assert replayed.manifest.rows_digest == manifest["rows_digest"]


def test_写盘会拦截混入的标签列(tmp_path: Path) -> None:
    batch = generate_batch(seed=5, row_count=3)
    batch.rows[0]["是否违规"] = True

    with pytest.raises(ValueError, match="标签泄漏"):
        write_batch(batch, tmp_path)
