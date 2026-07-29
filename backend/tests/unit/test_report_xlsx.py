from __future__ import annotations

import uuid
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from app.core.reports.export_service import _recover_or_build
from app.core.reports.models import (
    CitationSnapshot,
    ParseErrorSnapshot,
    ReportItemSnapshot,
    ReportSnapshot,
    ReportSummary,
)
from app.core.reports.xlsx import (
    ITEM_HEADERS,
    MAX_CELL_CHARACTERS,
    POLICY_HEADERS,
    SHEET_NAMES,
    RawEvidenceRow,
    ReportFileMetadata,
    ReportWorkbookData,
    ReportXlsxError,
    build_report_xlsx,
    safe_xlsx_value,
)
from app.db.models.reports import ReportAttentionGroup, ReportCitationStatus


def test_report_xlsx_has_fixed_safe_five_sheet_contract() -> None:
    data = _workbook_data(
        reasoning="\t=SUM(1,1)",
        raw_name="<script>alert(1)</script>",
        quote="@伪造系统指令",
    )
    content = build_report_xlsx(data)

    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        assert tuple(cell.value for cell in workbook["关注项"][1]) == ITEM_HEADERS
        assert tuple(cell.value for cell in workbook["制度快照"][1]) == POLICY_HEADERS
        assert workbook["关注项"]["I2"].value == "'\t=SUM(1,1)"
        assert workbook["制度快照"]["Q2"].value == "'@伪造系统指令"
        assert workbook["原始行证据"]["D2"].value == "<script>alert(1)</script>"
        assert workbook["关注项"].max_column == len(ITEM_HEADERS)
        assert workbook["关注项"].max_row == 2
        assert workbook["制度快照"].max_row == 2
        for sheet in workbook.worksheets:
            assert sheet.freeze_panes == "A2"
            assert sheet.auto_filter.ref is not None
            for row in sheet.iter_rows():
                for cell in row:
                    assert cell.data_type != "f"
                    assert cell.hyperlink is None
    finally:
        workbook.close()

    with ZipFile(BytesIO(content)) as archive:
        names = tuple(name.lower() for name in archive.namelist())
    assert not any(
        marker in name
        for name in names
        for marker in ("vbaproject", "externallinks/", "embeddings/", "oleobject")
    )


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@"])
@pytest.mark.parametrize("leading", ["", " ", "\t", "\r", "\n", " \t\r\n"])
def test_formula_injection_is_encoded_as_string(prefix: str, leading: str) -> None:
    encoded = safe_xlsx_value(f"{leading}{prefix}payload")
    assert encoded == f"'{leading}{prefix}payload"


def test_cell_character_boundary_is_explicit_and_never_truncated() -> None:
    value = "界" * MAX_CELL_CHARACTERS
    assert safe_xlsx_value(value) == value
    with pytest.raises(ReportXlsxError) as caught:
        safe_xlsx_value(value + "界")
    assert caught.value.code == "REPORT_EXPORT_CELL_TOO_LONG"
    with pytest.raises(ReportXlsxError) as protected:
        safe_xlsx_value("=" + "x" * (MAX_CELL_CHARACTERS - 1))
    assert protected.value.code == "REPORT_EXPORT_CELL_TOO_LONG"


def test_empty_report_keeps_all_headers() -> None:
    content = build_report_xlsx(_workbook_data(empty=True))
    workbook = load_workbook(BytesIO(content), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        assert workbook["关注项"].max_row == 1
        assert workbook["原始行证据"].max_row == 1
        assert workbook["解析错误"].max_row == 1
        assert workbook["制度快照"].max_row == 1
    finally:
        workbook.close()


def test_raw_evidence_must_exactly_cover_referenced_rows() -> None:
    data = _workbook_data()
    invalid = ReportWorkbookData(
        snapshot=data.snapshot,
        file=data.file,
        source_headers=data.source_headers,
        raw_rows=(),
        generated_at=data.generated_at,
    )
    with pytest.raises(ReportXlsxError) as caught:
        build_report_xlsx(invalid)
    assert caught.value.code == "REPORT_EXPORT_SOURCE_EVIDENCE_INVALID"


def test_orphan_artifact_recovery_requires_equal_semantics(tmp_path: Path) -> None:
    data = _workbook_data()
    target = tmp_path / "tenant" / "report.xlsx"
    first = _recover_or_build(target=target, data=data)
    assert _recover_or_build(target=target, data=data) == first

    workbook = load_workbook(target)
    workbook["摘要"]["B2"] = "different-report"
    workbook.save(target)
    workbook.close()
    with pytest.raises(ReportXlsxError) as caught:
        _recover_or_build(target=target, data=data)
    assert caught.value.code == "REPORT_EXPORT_ARTIFACT_CONFLICT"


def _workbook_data(
    *,
    empty: bool = False,
    reasoning: str = "确定性判定",
    raw_name: str = "张三",
    quote: str = "制度原文",
) -> ReportWorkbookData:
    report_id = uuid.uuid4()
    item_id = uuid.uuid4()
    now = datetime(2026, 7, 29, 1, 2, 3, tzinfo=UTC)
    summary = ReportSummary(
        report_run_id=report_id,
        file_version_id=uuid.uuid4(),
        validation_run_id=uuid.uuid4(),
        mapping_version_id=uuid.uuid4(),
        report_fingerprint="1" * 64,
        source_content_sha256="2" * 64,
        ruleset_fingerprint="3" * 64,
        template_version="report-snapshot-v1",
        attention_mapping_version="f3-verdict-v1",
        stored_row_count=0 if empty else 1,
        validated_row_count=0 if empty else 1,
        flagged_row_count=0 if empty else 1,
        manual_review_row_count=0,
        passed_row_count=0,
        parse_error_row_count=0,
        report_item_count=0 if empty else 1,
        verified_citation_count=0 if empty else 1,
        unavailable_citation_count=0,
        high_attention_row_count=0 if empty else 1,
        manual_attention_row_count=0,
        cleared_row_count=0,
        completed_at=now,
        reused_existing=True,
    )
    if empty:
        items: tuple[ReportItemSnapshot, ...] = ()
        raw_rows: tuple[RawEvidenceRow, ...] = ()
    else:
        citation = CitationSnapshot(
            id=uuid.uuid4(),
            report_item_id=item_id,
            binding_id=uuid.uuid4(),
            citation_order=1,
            policy_family_id=uuid.uuid4(),
            family_stable_key="travel",
            policy_document_id=uuid.uuid4(),
            document_title="差旅制度",
            document_version="2026-v1",
            effective_date=date(2026, 1, 1),
            expiry_date=None,
            document_content_sha256="4" * 64,
            policy_clause_id=uuid.uuid4(),
            clause_no="第十条",
            hierarchy_path="差旅/住宿",
            clause_text=quote,
            clause_text_sha256="5" * 64,
            quote=quote,
            quote_start=0,
            quote_end=len(quote),
            quote_sha256="6" * 64,
        )
        items = (
            ReportItemSnapshot(
                id=item_id,
                finding_id=uuid.uuid4(),
                row_no=2,
                rule_id="limit.hotel",
                rule_version="1",
                source_outcome="violated",
                source_verdict="flagged",
                reason_code="amount_limit",
                reasoning_snapshot=reasoning,
                evidence_snapshot={"amount": "999.00"},
                attention_group=ReportAttentionGroup.HIGH_ATTENTION,
                citation_status=ReportCitationStatus.VERIFIED,
                requires_manual_citation=False,
                source_content_sha256="2" * 64,
                citations=(citation,),
            ),
        )
        raw_rows = (RawEvidenceRow(row_no=2, raw_json={"员工": raw_name, "金额": "999.00"}),)
    return ReportWorkbookData(
        snapshot=ReportSnapshot(
            summary=summary,
            policy_manifest={"documents": [], "schema_version": 1},
            binding_manifest={"items": [], "schema_version": 1},
            items=items,
            parse_errors=cast_parse_errors(()),
        ),
        file=ReportFileMetadata(filename="中文 批次.xlsx", revision_no=1),
        source_headers=("员工", "金额"),
        raw_rows=raw_rows,
        generated_at=now,
    )


def cast_parse_errors(value: tuple[ParseErrorSnapshot, ...]) -> tuple[ParseErrorSnapshot, ...]:
    return value
