from __future__ import annotations

import hashlib
import uuid
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.errors import NotFoundError
from app.core.reports.export_service import (
    ReportExportError,
    create_report_export,
    download_report_export,
)
from app.core.reports.service import generate_report
from app.core.tenancy.scope import bind_tenant
from app.db.models.audit import AuditLog
from app.db.models.batch import FileVersion
from app.db.models.reports import ReportExport
from tests.integration.test_report_service import (
    _seed_policy_bindings,
    _seed_tenant_only,
    _seed_validated_batch,
)

pytestmark = pytest.mark.integration


async def test_xlsx_export_replay_download_audit_and_hash_fail_closed(
    session_factory: async_sessionmaker[AsyncSession], tmp_path: Path
) -> None:
    slug = f"xlsx-export-{uuid.uuid4().hex[:8]}"
    tenant_id, actor_id, batch_id = await _seed_validated_batch(session_factory, slug=slug)
    await _seed_policy_bindings(
        session_factory,
        tenant_id=tenant_id,
        actor_id=actor_id,
        batch_id=batch_id,
    )
    upload_root = tmp_path / "uploads"
    export_root = tmp_path / "exports"
    source = _source_workbook()
    source_hash = hashlib.sha256(source).hexdigest()
    source_path = upload_root / str(tenant_id) / f"{source_hash}.xlsx"
    source_path.parent.mkdir(parents=True)
    source_path.write_bytes(source)
    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        await session.execute(
            update(FileVersion)
            .where(FileVersion.id == batch_id)
            .values(content_hash=source_hash, filename="中文批次.xlsx")
        )
        await session.commit()

    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        report = await generate_report(
            session,
            session_factory,
            file_version_id=batch_id,
            tenant_id=tenant_id,
            actor_id=actor_id,
            idempotency_key="xlsx-report-key",
        )
        await session.commit()

    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        created = await create_report_export(
            session,
            session_factory,
            tenant_id=tenant_id,
            actor_id=actor_id,
            report_run_id=report.report_run_id,
            idempotency_key="xlsx-export-key",
            export_root=export_root,
            upload_root=upload_root,
        )
        await session.commit()
    assert created.reused_existing is False

    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        replayed = await create_report_export(
            session,
            session_factory,
            tenant_id=tenant_id,
            actor_id=actor_id,
            report_run_id=report.report_run_id,
            idempotency_key="xlsx-export-key",
            export_root=export_root,
            upload_root=upload_root,
        )
        first_download = await download_report_export(
            session,
            tenant_id=tenant_id,
            actor_id=actor_id,
            export_id=created.export_id,
            export_root=export_root,
        )
        second_download = await download_report_export(
            session,
            tenant_id=tenant_id,
            actor_id=actor_id,
            export_id=created.export_id,
            export_root=export_root,
        )
        await session.commit()
    assert replayed.reused_existing is True
    assert replayed.artifact_sha256 == created.artifact_sha256
    assert first_download.content == second_download.content
    assert first_download.filename == "费用预审报告-中文批次.xlsx"
    workbook = load_workbook(BytesIO(first_download.content), read_only=True)
    assert workbook.sheetnames == ["摘要", "关注项", "原始行证据", "解析错误", "制度快照"]
    workbook.close()

    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        generate_count = int(
            await session.scalar(
                select(func.count())
                .select_from(AuditLog)
                .where(AuditLog.action == "report.export_generate")
            )
            or 0
        )
        download_count = int(
            await session.scalar(
                select(func.count())
                .select_from(AuditLog)
                .where(AuditLog.action == "report.export_download")
            )
            or 0
        )
        export = await session.get(ReportExport, created.export_id)
        assert export is not None and export.artifact_storage_key is not None
    assert generate_count == 1
    assert download_count == 2

    other_tenant, other_actor, _ = await _seed_tenant_only(
        session_factory, slug=f"xlsx-export-other-{uuid.uuid4().hex[:8]}"
    )
    async with session_factory() as session:
        bind_tenant(session.sync_session, other_tenant)
        with pytest.raises(NotFoundError) as isolated:
            await download_report_export(
                session,
                tenant_id=other_tenant,
                actor_id=other_actor,
                export_id=created.export_id,
                export_root=export_root,
            )
        assert getattr(isolated.value, "code", None) == "REPORT_EXPORT_NOT_FOUND"
        await session.rollback()

    artifact = export_root / export.artifact_storage_key
    artifact.write_bytes(b"tampered")
    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        with pytest.raises(ReportExportError) as caught:
            await download_report_export(
                session,
                tenant_id=tenant_id,
                actor_id=actor_id,
                export_id=created.export_id,
                export_root=export_root,
            )
        assert caught.value.code == "REPORT_EXPORT_ARTIFACT_HASH_MISMATCH"
        await session.rollback()


async def test_xlsx_export_failure_is_persisted_without_source_text(
    session_factory: async_sessionmaker[AsyncSession], tmp_path: Path
) -> None:
    slug = f"xlsx-export-failure-{uuid.uuid4().hex[:8]}"
    tenant_id, actor_id, batch_id = await _seed_validated_batch(session_factory, slug=slug)
    await _seed_policy_bindings(
        session_factory,
        tenant_id=tenant_id,
        actor_id=actor_id,
        batch_id=batch_id,
    )
    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        report = await generate_report(
            session,
            session_factory,
            file_version_id=batch_id,
            tenant_id=tenant_id,
            actor_id=actor_id,
            idempotency_key="xlsx-failure-report-key",
        )
        await session.commit()

    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        with pytest.raises(ReportExportError) as caught:
            await create_report_export(
                session,
                session_factory,
                tenant_id=tenant_id,
                actor_id=actor_id,
                report_run_id=report.report_run_id,
                idempotency_key="xlsx-failure-export-key",
                export_root=tmp_path / "exports",
                upload_root=tmp_path / "missing-uploads",
            )
        assert caught.value.code == "REPORT_EXPORT_SOURCE_UNAVAILABLE"

    async with session_factory() as session:
        bind_tenant(session.sync_session, tenant_id)
        export = await session.scalar(
            select(ReportExport).where(ReportExport.report_run_id == report.report_run_id)
        )
        assert export is not None
        assert export.status == "failed"
        assert export.failure_code == "REPORT_EXPORT_SOURCE_UNAVAILABLE"
        audit = await session.scalar(
            select(AuditLog).where(AuditLog.action == "report.export_failed")
        )
        assert audit is not None
        assert set(audit.payload_json or {}) == {"report_run_id", "failure_code"}


def _source_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["row"])
    sheet.append([2])
    sheet.append([3])
    sheet.append([4])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
